from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List

import networkx as nx
import numpy as np
from openpyxl import load_workbook
from sklearn.cluster import SpectralClustering


class DataValidationError(ValueError):
    """Raised when benchmark data violates model assumptions."""


@dataclass(frozen=True, slots=True)
class CustomerNode:
    """Customer node.

    For node \(i\), fields are demand \(q_i\), ready time \(a_i\), due time
    \(b_i\), and service time \(s_i\). Compatibility preprocessing assumes
    service at node \(i\) starts at \(a_i\), so arrival at node \(j\) is
    \(t_j=a_i+s_i+d_{ij}\).
    """

    id: int
    demand: float
    ready_time: float
    due_time: float
    service_time: float

    def __post_init__(self) -> None:
        if self.id < 0:
            raise DataValidationError(f"Invalid node id: {self.id}.")
        if min(self.demand, self.ready_time, self.due_time, self.service_time) < 0:
            raise DataValidationError(f"Node {self.id} contains negative attributes.")
        if self.ready_time > self.due_time:
            raise DataValidationError(f"Node {self.id} has ready_time > due_time.")


@dataclass(frozen=True, slots=True)
class LogisticsInstance:
    """Immutable instance loaded from Excel.

    `all_nodes` is ordered by node id and `travel_time_matrix[i, j] = D_{ij}`.
    The official file has 51 nodes including depot 0, so customer-only matrices
    have shape \(50\times 50\).
    """

    depot: CustomerNode
    customers: List[CustomerNode]
    all_nodes: List[CustomerNode]
    travel_time_matrix: np.ndarray
    capacity: float

    def __post_init__(self) -> None:
        n = len(self.all_nodes)
        if self.depot.id != 0:
            raise DataValidationError("Depot id must be 0.")
        if self.travel_time_matrix.shape != (n, n):
            raise DataValidationError(f"Expected matrix shape {(n, n)}, got {self.travel_time_matrix.shape}.")
        if np.any(self.travel_time_matrix < 0):
            raise DataValidationError("Travel matrix cannot contain negative values.")
        if self.capacity <= 0:
            raise DataValidationError("Vehicle capacity must be positive.")
        ids = [node.id for node in self.all_nodes]
        if ids != list(range(n)):
            raise DataValidationError(f"Node ids must be contiguous from 0, got {ids}.")

    @property
    def customer_count(self) -> int:
        return len(self.customers)

    @property
    def customer_ids(self) -> List[int]:
        return [node.id for node in self.customers]

    def get_node(self, node_id: int) -> CustomerNode:
        if node_id < 0 or node_id >= len(self.all_nodes):
            raise IndexError(f"Node id {node_id} out of range.")
        return self.all_nodes[node_id]


class DataLoader:
    """Read the official Excel benchmark.

    Sheet 1 is parsed by position:
    `node_id, ready_time, due_time, service_time, demand, unused, capacity`.
    Sheet 2 is the full transportation-time matrix with row/column ids.
    This avoids localized-header encoding issues.
    """

    EXPECTED_NODE_COUNT: int = 51

    def __init__(self, excel_path: str | Path) -> None:
        self._excel_path = Path(excel_path)

    def load(self) -> LogisticsInstance:
        if not self._excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self._excel_path}")
        workbook = load_workbook(self._excel_path, data_only=True)
        if len(workbook.sheetnames) < 2:
            raise DataValidationError("Workbook must contain at least two sheets.")
        nodes, capacity = self._parse_nodes(workbook[workbook.sheetnames[0]])
        matrix = self._parse_matrix(workbook[workbook.sheetnames[1]], len(nodes))
        return LogisticsInstance(nodes[0], list(nodes[1:]), list(nodes), matrix, capacity)

    def _parse_nodes(self, sheet: object) -> tuple[List[CustomerNode], float]:
        rows = [r for r in sheet.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
        if not rows:
            raise DataValidationError("Attribute sheet is empty.")
        nodes: List[CustomerNode] = []
        capacity: float | None = None
        for row_idx, row in enumerate(rows, start=2):
            node = CustomerNode(
                id=self._as_int(row[0], f"row {row_idx} node_id"),
                ready_time=self._as_float(row[1], f"row {row_idx} ready_time"),
                due_time=self._as_float(row[2], f"row {row_idx} due_time"),
                service_time=self._as_float(row[3], f"row {row_idx} service_time"),
                demand=self._as_float(row[4], f"row {row_idx} demand"),
            )
            nodes.append(node)
            if len(row) > 6 and row[6] is not None:
                current = self._as_float(row[6], f"row {row_idx} capacity")
                if capacity is None:
                    capacity = current
                elif not np.isclose(capacity, current):
                    raise DataValidationError("Capacity column must be constant.")
        ids = [node.id for node in nodes]
        if ids != list(range(len(nodes))):
            raise DataValidationError(f"Node ids must be contiguous from 0, got {ids}.")
        if len(nodes) != self.EXPECTED_NODE_COUNT:
            raise DataValidationError(f"Expected {self.EXPECTED_NODE_COUNT} nodes, got {len(nodes)}.")
        if capacity is None:
            raise DataValidationError("Capacity column is missing.")
        return nodes, capacity

    def _parse_matrix(self, sheet: object, node_count: int) -> np.ndarray:
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header is None:
            raise DataValidationError("Travel-time sheet is empty.")
        header_ids = [self._as_int(v, f"header col {i}") for i, v in enumerate(header[1:], start=1) if v is not None]
        expected = list(range(node_count))
        if header_ids != expected:
            raise DataValidationError(f"Header ids mismatch: expected {expected}, got {header_ids}.")
        rows = [r for r in sheet.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
        if len(rows) != node_count:
            raise DataValidationError(f"Expected {node_count} matrix rows, got {len(rows)}.")
        matrix = np.full((node_count, node_count), np.inf, dtype=np.float64)
        for row_idx, row in enumerate(rows):
            row_id = self._as_int(row[0], f"matrix row {row_idx + 2} id")
            if row_id != row_idx:
                raise DataValidationError(f"Matrix row order mismatch: expected {row_idx}, got {row_id}.")
            if len(row) < node_count + 1:
                raise DataValidationError(f"Matrix row {row_idx + 2} has too few columns.")
            for col_idx in range(node_count):
                value = self._as_float(row[col_idx + 1], f"D[{row_id},{col_idx}]")
                if value < 0:
                    raise DataValidationError(f"D[{row_id},{col_idx}] cannot be negative.")
                matrix[row_id, col_idx] = value
        return matrix

    @staticmethod
    def _as_int(value: object, name: str) -> int:
        if value is None:
            raise DataValidationError(f"Missing value for {name}.")
        try:
            return int(value)
        except (TypeError, ValueError) as exc:
            raise DataValidationError(f"Invalid integer for {name}: {value!r}") from exc

    @staticmethod
    def _as_float(value: object, name: str) -> float:
        if value is None:
            raise DataValidationError(f"Missing value for {name}.")
        try:
            numeric = float(value)
        except (TypeError, ValueError) as exc:
            raise DataValidationError(f"Invalid float for {name}: {value!r}") from exc
        if np.isnan(numeric):
            raise DataValidationError(f"NaN is not allowed for {name}.")
        return numeric


@dataclass(frozen=True, slots=True)
class CompatibilityComputationResult:
    """Bundle containing penalty matrix \(P\) and compatibility matrix \(W\)."""

    alpha: float
    beta: float
    epsilon: float
    penalty_matrix: np.ndarray
    compatibility_matrix: np.ndarray


class TemporalCompatibilityGraph:
    """Construct the directed spatio-temporal compatibility graph.

    For customers \(i\neq j\), with earliest-start departure from \(i\),
    \(t_j = a_i + s_i + d_{ij}\).

    The inevitable penalty is
    \(P_{ij}=10\max(a_j-t_j,0)^2 + 20\max(t_j-b_j,0)^2\).

    The compatibility score is
    \(W_{ij}=\alpha/(D_{ij}+\varepsilon)-\beta P_{ij}\).

    Distinct-node zero distances are stabilized by `distance_floor`, and infinite
    distances produce `W_ij = -inf`.
    """

    EARLY_COEF: float = 10.0
    LATE_COEF: float = 20.0

    def __init__(self, instance: LogisticsInstance, alpha: float, beta: float, epsilon: float = 1e-6, distance_floor: float = 1e-9) -> None:
        self._instance = instance
        self._alpha = float(alpha)
        self._beta = float(beta)
        self._epsilon = float(epsilon)
        self._distance_floor = float(distance_floor)
        self._validate_params()

    def compute_mandatory_departure_penalty(self, from_node_id: int, to_node_id: int) -> float:
        self._validate_customer_id(from_node_id)
        self._validate_customer_id(to_node_id)
        if from_node_id == to_node_id:
            return 0.0
        source = self._instance.get_node(from_node_id)
        target = self._instance.get_node(to_node_id)
        distance = self._instance.travel_time_matrix[from_node_id, to_node_id]
        if np.isinf(distance):
            return float("inf")
        arrival = source.ready_time + source.service_time + distance
        early = max(target.ready_time - arrival, 0.0)
        late = max(arrival - target.due_time, 0.0)
        return self.EARLY_COEF * early * early + self.LATE_COEF * late * late

    def compute_pairwise_compatibility(self, from_node_id: int, to_node_id: int) -> float:
        self._validate_customer_id(from_node_id)
        self._validate_customer_id(to_node_id)
        if from_node_id == to_node_id:
            return 0.0
        distance = self._instance.travel_time_matrix[from_node_id, to_node_id]
        penalty = self.compute_mandatory_departure_penalty(from_node_id, to_node_id)
        if np.isinf(distance) or np.isinf(penalty):
            return float("-inf")
        denom = max(distance, self._distance_floor) + self._epsilon
        return self._alpha / denom - self._beta * penalty

    def build_penalty_matrix(self) -> np.ndarray:
        ids = self._instance.customer_ids
        out = np.zeros((len(ids), len(ids)), dtype=np.float64)
        for r, i in enumerate(ids):
            for c, j in enumerate(ids):
                out[r, c] = self.compute_mandatory_departure_penalty(i, j)
        return out

    def build_compatibility_matrix(self) -> np.ndarray:
        ids = self._instance.customer_ids
        out = np.zeros((len(ids), len(ids)), dtype=np.float64)
        for r, i in enumerate(ids):
            for c, j in enumerate(ids):
                out[r, c] = self.compute_pairwise_compatibility(i, j)
        return out

    def build(self) -> CompatibilityComputationResult:
        p = self.build_penalty_matrix()
        w = self.build_compatibility_matrix()
        return CompatibilityComputationResult(self._alpha, self._beta, self._epsilon, p, w)

    def to_networkx_digraph(self, compatibility_matrix: np.ndarray | None = None) -> nx.DiGraph:
        matrix = compatibility_matrix if compatibility_matrix is not None else self.build_compatibility_matrix()
        expected = (self._instance.customer_count, self._instance.customer_count)
        if matrix.shape != expected:
            raise DataValidationError(f"Compatibility matrix must have shape {expected}.")
        graph = nx.DiGraph()
        for node in self._instance.customers:
            graph.add_node(node.id, demand=node.demand, ready_time=node.ready_time, due_time=node.due_time, service_time=node.service_time)
        ids = self._instance.customer_ids
        for r, i in enumerate(ids):
            for c, j in enumerate(ids):
                if i != j:
                    graph.add_edge(i, j, weight=float(matrix[r, c]))
        return graph

    def spectral_cluster(self, n_clusters: int, compatibility_matrix: np.ndarray | None = None, random_state: int = 42) -> np.ndarray:
        """Cluster customers using the non-negative affinity surrogate `max(W, 0)`."""
        if n_clusters <= 0:
            raise DataValidationError("n_clusters must be positive.")
        matrix = compatibility_matrix if compatibility_matrix is not None else self.build_compatibility_matrix()
        expected = (self._instance.customer_count, self._instance.customer_count)
        if matrix.shape != expected:
            raise DataValidationError(f"Compatibility matrix must have shape {expected}.")
        affinity = np.maximum(matrix, 0.0)
        np.fill_diagonal(affinity, 1.0)
        model = SpectralClustering(n_clusters=n_clusters, affinity="precomputed", random_state=random_state, assign_labels="kmeans")
        return model.fit_predict(affinity)

    def export_customer_feature_matrix(self) -> np.ndarray:
        return np.asarray([[n.demand, n.ready_time, n.due_time, n.service_time] for n in self._instance.customers], dtype=np.float64)

    def _validate_params(self) -> None:
        if not np.isfinite(self._alpha) or self._alpha < 0:
            raise DataValidationError("alpha must be finite and non-negative.")
        if not np.isfinite(self._beta) or self._beta < 0:
            raise DataValidationError("beta must be finite and non-negative.")
        if not np.isfinite(self._epsilon) or self._epsilon <= 0:
            raise DataValidationError("epsilon must be finite and positive.")
        if not np.isfinite(self._distance_floor) or self._distance_floor <= 0:
            raise DataValidationError("distance_floor must be finite and positive.")

    def _validate_customer_id(self, node_id: int) -> None:
        if node_id == 0:
            raise DataValidationError("Compatibility graph excludes depot 0.")
        if node_id < 1 or node_id > self._instance.customer_count:
            raise DataValidationError(f"Customer id must be in [1, {self._instance.customer_count}], got {node_id}.")


__all__: List[str] = ["CompatibilityComputationResult", "CustomerNode", "DataLoader", "DataValidationError", "LogisticsInstance", "TemporalCompatibilityGraph"]
