from __future__ import annotations

import argparse
import json
import random
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

from openpyxl import load_workbook


EARLY_PENALTY = 10
LATE_PENALTY = 20
BIG_M = 1_000_000


@dataclass
class Instance:
    customers: List[int]
    travel: Dict[int, Dict[int, int]]
    lower: Dict[int, int]
    upper: Dict[int, int]
    service: Dict[int, int]
    demand: Dict[int, int]
    capacity: int


@dataclass
class RouteMetrics:
    route: List[int]
    total_travel: int
    total_penalty: int
    objective: int
    load: int
    start_times: Dict[int, int]
    early_violation: Dict[int, int]
    late_violation: Dict[int, int]
    penalties: Dict[int, int]


def load_instance(path: Path) -> Instance:
    workbook = load_workbook(path, data_only=True)
    attr_sheet = workbook[workbook.sheetnames[0]]
    matrix_sheet = workbook[workbook.sheetnames[1]]

    rows = list(attr_sheet.iter_rows(min_row=2, values_only=True))
    capacity = int(rows[0][6])

    lower: Dict[int, int] = {}
    upper: Dict[int, int] = {}
    service: Dict[int, int] = {}
    demand: Dict[int, int] = {}

    for row in rows:
        node = int(row[0])
        lower[node] = int(row[1])
        upper[node] = int(row[2])
        service[node] = int(row[3])
        demand[node] = int(row[4])

    header = [int(value) for value in next(matrix_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[1:]]
    travel: Dict[int, Dict[int, int]] = {}
    for row in matrix_sheet.iter_rows(min_row=2, values_only=True):
        node = int(row[0])
        travel[node] = {header[idx]: int(row[idx + 1]) for idx in range(len(header))}

    return Instance(
        customers=list(range(1, len(rows))),
        travel=travel,
        lower=lower,
        upper=upper,
        service=service,
        demand=demand,
        capacity=capacity,
    )


def time_window_penalty(instance: Instance, node: int, start_time: int) -> int:
    early = max(instance.lower[node] - start_time, 0)
    late = max(start_time - instance.upper[node], 0)
    return EARLY_PENALTY * early * early + LATE_PENALTY * late * late


def evaluate_route(instance: Instance, route: Sequence[int]) -> RouteMetrics:
    current_time = 0
    last = 0
    total_travel = 0
    total_penalty = 0
    load = 0
    start_times: Dict[int, int] = {}
    early_violation: Dict[int, int] = {}
    late_violation: Dict[int, int] = {}
    penalties: Dict[int, int] = {}

    for node in route:
        total_travel += instance.travel[last][node]
        current_time += instance.travel[last][node]
        load += instance.demand[node]

        start_times[node] = current_time
        early_violation[node] = max(instance.lower[node] - current_time, 0)
        late_violation[node] = max(current_time - instance.upper[node], 0)
        penalties[node] = time_window_penalty(instance, node, current_time)
        total_penalty += penalties[node]

        current_time += instance.service[node]
        last = node

    total_travel += instance.travel[last][0]
    return RouteMetrics(
        route=list(route),
        total_travel=total_travel,
        total_penalty=total_penalty,
        objective=total_travel + total_penalty,
        load=load,
        start_times=start_times,
        early_violation=early_violation,
        late_violation=late_violation,
        penalties=penalties,
    )


def route_summary(metrics: RouteMetrics) -> Dict[str, object]:
    ordered_starts = [{"node": node, "start_time": metrics.start_times[node]} for node in metrics.route]
    ordered_violation = [
        {
            "node": node,
            "early": metrics.early_violation[node],
            "late": metrics.late_violation[node],
            "penalty": metrics.penalties[node],
        }
        for node in metrics.route
    ]
    return {
        "route": [0] + metrics.route + [0],
        "load": metrics.load,
        "travel_time": metrics.total_travel,
        "time_window_penalty": metrics.total_penalty,
        "objective": metrics.objective,
        "starts": ordered_starts,
        "violations": ordered_violation,
    }


def held_karp_tsp(instance: Instance, customers: Sequence[int]) -> RouteMetrics:
    nodes = list(customers)
    n = len(nodes)
    full_mask = (1 << n) - 1
    inf = 10**18

    dp = [[inf] * n for _ in range(1 << n)]
    parent = [[-1] * n for _ in range(1 << n)]

    for j, node in enumerate(nodes):
        dp[1 << j][j] = instance.travel[0][node]

    for mask in range(1 << n):
        remaining = full_mask ^ mask
        for j, node_j in enumerate(nodes):
            if not (mask & (1 << j)):
                continue
            current = dp[mask][j]
            if current >= inf:
                continue
            rem = remaining
            while rem:
                bit = rem & -rem
                k = bit.bit_length() - 1
                node_k = nodes[k]
                next_mask = mask | bit
                candidate = current + instance.travel[node_j][node_k]
                if candidate < dp[next_mask][k]:
                    dp[next_mask][k] = candidate
                    parent[next_mask][k] = j
                rem ^= bit

    best_cost = inf
    last_index = -1
    for j, node in enumerate(nodes):
        candidate = dp[full_mask][j] + instance.travel[node][0]
        if candidate < best_cost:
            best_cost = candidate
            last_index = j

    route: List[int] = []
    mask = full_mask
    while last_index != -1:
        route.append(nodes[last_index])
        prev = parent[mask][last_index]
        mask ^= 1 << last_index
        last_index = prev
    route.reverse()
    return evaluate_route(instance, route)


def exact_soft_time_window_tsp(instance: Instance, customers: Sequence[int]) -> RouteMetrics:
    nodes = list(customers)
    n = len(nodes)
    full_mask = (1 << n) - 1

    service_sum = [0] * (1 << n)
    for mask in range(1, 1 << n):
        bit = mask & -mask
        idx = bit.bit_length() - 1
        service_sum[mask] = service_sum[mask ^ bit] + instance.service[nodes[idx]]

    # Keep all unique (subset, last, travel_prefix) states.
    # Earlier arrival is not always better here because it may trigger larger early penalties later.
    states: List[Dict[int, Dict[int, int]]] = [dict() for _ in range(1 << n)]
    parents: Dict[tuple[int, int, int], tuple[int, int, int] | None] = {}

    for idx, node in enumerate(nodes):
        mask = 1 << idx
        travel_prefix = instance.travel[0][node]
        penalty = time_window_penalty(instance, node, travel_prefix)
        states[mask][idx] = {travel_prefix: penalty}
        parents[(mask, idx, travel_prefix)] = None

    for size in range(1, n):
        for mask in range(1 << n):
            if mask.bit_count() != size:
                continue
            label_sets = states[mask]
            if not label_sets:
                continue
            remaining = full_mask ^ mask
            visited_service = service_sum[mask]
            for last_idx, label_map in label_sets.items():
                last_node = nodes[last_idx]
                for travel_prefix, accumulated_penalty in label_map.items():
                    elapsed_after_service = travel_prefix + visited_service
                    rem = remaining
                    while rem:
                        bit = rem & -rem
                        next_idx = bit.bit_length() - 1
                        next_node = nodes[next_idx]
                        next_travel_prefix = travel_prefix + instance.travel[last_node][next_node]
                        next_start = elapsed_after_service + instance.travel[last_node][next_node]
                        next_penalty = accumulated_penalty + time_window_penalty(instance, next_node, next_start)
                        next_mask = mask | bit
                        next_labels = states[next_mask].setdefault(next_idx, {})
                        old_penalty = next_labels.get(next_travel_prefix)
                        if old_penalty is None or next_penalty < old_penalty:
                            next_labels[next_travel_prefix] = next_penalty
                            parents[(next_mask, next_idx, next_travel_prefix)] = (mask, last_idx, travel_prefix)
                        rem ^= bit

    best_objective = None
    best_state = None
    for last_idx, label_map in states[full_mask].items():
        last_node = nodes[last_idx]
        for travel_prefix, accumulated_penalty in label_map.items():
            objective = travel_prefix + instance.travel[last_node][0] + accumulated_penalty
            if best_objective is None or objective < best_objective:
                best_objective = objective
                best_state = (full_mask, last_idx, travel_prefix)

    if best_state is None:
        raise ValueError("Exact DP for problem 2 failed to produce a route.")

    route: List[int] = []
    state = best_state
    while state is not None:
        mask, last_idx, travel_prefix = state
        route.append(nodes[last_idx])
        state = parents[(mask, last_idx, travel_prefix)]
    route.reverse()
    return evaluate_route(instance, route)


def best_insertion_route(instance: Instance, order: Sequence[int]) -> List[int]:
    route: List[int] = []
    for node in order:
        best_route = None
        best_obj = None
        for pos in range(len(route) + 1):
            candidate = route[:pos] + [node] + route[pos:]
            objective = evaluate_route(instance, candidate).objective
            if best_obj is None or objective < best_obj:
                best_obj = objective
                best_route = candidate
        route = best_route if best_route is not None else [node]
    return route


def local_search_single(
    instance: Instance,
    route: Sequence[int],
    max_rounds: int = 3,
    enable_reverse: bool = False,
) -> List[int]:
    current = list(route)
    for _ in range(max_rounds):
        base_objective = evaluate_route(instance, current).objective
        best_objective = base_objective
        best_route = current
        n = len(current)

        for i in range(n):
            for j in range(n + 1):
                if j == i or j == i + 1:
                    continue
                candidate = current[:]
                node = candidate.pop(i)
                insert_at = j - 1 if j > i else j
                candidate.insert(insert_at, node)
                objective = evaluate_route(instance, candidate).objective
                if objective < best_objective:
                    best_objective = objective
                    best_route = candidate

        for i in range(n):
            for j in range(i + 1, n):
                candidate = current[:]
                candidate[i], candidate[j] = candidate[j], candidate[i]
                objective = evaluate_route(instance, candidate).objective
                if objective < best_objective:
                    best_objective = objective
                    best_route = candidate

        if enable_reverse:
            for i in range(n):
                for j in range(i + 1, n):
                    candidate = current[:i] + list(reversed(current[i : j + 1])) + current[j + 1 :]
                    objective = evaluate_route(instance, candidate).objective
                    if objective < best_objective:
                        best_objective = objective
                        best_route = candidate

        if best_objective >= base_objective:
            break
        current = best_route
    return current


def candidate_orders(instance: Instance, customers: Sequence[int], rng: random.Random) -> List[List[int]]:
    customers = list(customers)
    orders = [
        sorted(customers, key=lambda node: (instance.upper[node], instance.lower[node], instance.travel[0][node])),
        sorted(customers, key=lambda node: (instance.upper[node] - instance.lower[node], instance.upper[node], instance.travel[0][node])),
        sorted(customers, key=lambda node: (instance.travel[0][node], instance.upper[node], instance.lower[node])),
        sorted(customers, key=lambda node: (instance.upper[node] + instance.travel[0][node], instance.lower[node])),
    ]
    for _ in range(6):
        order = customers[:]
        rng.shuffle(order)
        orders.append(order)
    return orders


def multi_start_single_vehicle(
    instance: Instance,
    customers: Sequence[int],
    extra_routes: Iterable[Sequence[int]] | None = None,
    seed: int = 0,
    max_rounds: int = 3,
    enable_reverse: bool = False,
) -> RouteMetrics:
    rng = random.Random(seed)
    best_metrics: RouteMetrics | None = None

    for order in candidate_orders(instance, customers, rng):
        route = best_insertion_route(instance, order)
        route = local_search_single(instance, route, max_rounds=max_rounds, enable_reverse=enable_reverse)
        metrics = evaluate_route(instance, route)
        if best_metrics is None or metrics.objective < best_metrics.objective:
            best_metrics = metrics

    if extra_routes:
        for route in extra_routes:
            improved = local_search_single(instance, route, max_rounds=max_rounds, enable_reverse=enable_reverse)
            metrics = evaluate_route(instance, improved)
            if best_metrics is None or metrics.objective < best_metrics.objective:
                best_metrics = metrics

    if best_metrics is None:
        raise ValueError("No route candidates were generated.")
    return best_metrics


def best_single_vehicle_over_seeds(
    instance: Instance,
    customers: Sequence[int],
    seeds: Iterable[int],
    extra_routes: Iterable[Sequence[int]] | None = None,
    max_rounds: int = 3,
    enable_reverse: bool = False,
) -> RouteMetrics:
    best_metrics: RouteMetrics | None = None
    for seed in seeds:
        metrics = multi_start_single_vehicle(
            instance,
            customers,
            extra_routes=extra_routes,
            seed=seed,
            max_rounds=max_rounds,
            enable_reverse=enable_reverse,
        )
        if best_metrics is None or metrics.objective < best_metrics.objective:
            best_metrics = metrics
    if best_metrics is None:
        raise ValueError("No route candidates were generated.")
    return best_metrics


def route_cost(instance: Instance, route: Sequence[int]) -> int:
    return evaluate_route(instance, route).objective if route else 0


def build_fixed_vehicle_solution(
    instance: Instance,
    customers: Sequence[int],
    vehicle_count: int,
) -> List[List[int]]:
    order = sorted(
        customers,
        key=lambda node: (instance.upper[node], instance.lower[node], -instance.demand[node], instance.travel[0][node]),
    )
    if vehicle_count > len(order):
        raise ValueError("Vehicle count cannot exceed customer count.")

    routes: List[List[int]] = [[node] for node in order[:vehicle_count]]
    loads = [instance.demand[node] for node in order[:vehicle_count]]

    for node in order[vehicle_count:]:
        best_vehicle = -1
        best_route: List[int] | None = None
        best_delta = None
        for vehicle in range(vehicle_count):
            if loads[vehicle] + instance.demand[node] > instance.capacity:
                continue
            base_cost = route_cost(instance, routes[vehicle])
            for pos in range(len(routes[vehicle]) + 1):
                candidate = routes[vehicle][:pos] + [node] + routes[vehicle][pos:]
                delta = route_cost(instance, candidate) - base_cost
                if best_delta is None or delta < best_delta:
                    best_delta = delta
                    best_vehicle = vehicle
                    best_route = candidate
        if best_vehicle == -1 or best_route is None:
            raise RuntimeError(f"Failed to place customer {node} with {vehicle_count} vehicles.")
        routes[best_vehicle] = best_route
        loads[best_vehicle] += instance.demand[node]
    return routes


def improve_multi_vehicle_routes(
    instance: Instance,
    routes: Sequence[Sequence[int]],
    keep_nonempty: bool = True,
    max_rounds: int = 8,
) -> List[List[int]]:
    current = [local_search_single(instance, route, max_rounds=2, enable_reverse=False) for route in routes]

    for _ in range(max_rounds):
        best_delta = 0
        best_move = None

        for from_idx, from_route in enumerate(current):
            if not from_route:
                continue
            base_from = route_cost(instance, from_route)
            from_load = evaluate_route(instance, from_route).load
            for to_idx, to_route in enumerate(current):
                if from_idx == to_idx:
                    continue
                base_to = route_cost(instance, to_route)
                to_load = evaluate_route(instance, to_route).load if to_route else 0
                for remove_pos, node in enumerate(from_route):
                    if keep_nonempty and len(from_route) == 1:
                        continue
                    if to_load + instance.demand[node] > instance.capacity:
                        continue
                    new_from = from_route[:remove_pos] + from_route[remove_pos + 1 :]
                    new_from_cost = route_cost(instance, new_from)
                    for insert_pos in range(len(to_route) + 1):
                        new_to = to_route[:insert_pos] + [node] + to_route[insert_pos:]
                        new_to_cost = route_cost(instance, new_to)
                        delta = (new_from_cost + new_to_cost) - (base_from + base_to)
                        if delta < best_delta:
                            best_delta = delta
                            best_move = (from_idx, to_idx, new_from, new_to)

        if best_move is None:
            break

        from_idx, to_idx, new_from, new_to = best_move
        current[from_idx] = local_search_single(instance, new_from, max_rounds=2, enable_reverse=False)
        current[to_idx] = local_search_single(instance, new_to, max_rounds=2, enable_reverse=False)
    return current


def evaluate_routes(instance: Instance, routes: Sequence[Sequence[int]]) -> Dict[str, object]:
    used_metrics = [evaluate_route(instance, route) for route in routes if route]
    travel_time = sum(metrics.total_travel for metrics in used_metrics)
    total_penalty = sum(metrics.total_penalty for metrics in used_metrics)
    used_vehicles = len(used_metrics)
    return {
        "vehicles_used": used_vehicles,
        "travel_time": travel_time,
        "time_window_penalty": total_penalty,
        "secondary_objective": travel_time + total_penalty,
        "hierarchical_objective": used_vehicles * BIG_M + travel_time + total_penalty,
        "routes": [route_summary(metrics) for metrics in used_metrics],
    }


def solve_problem_1(instance: Instance) -> Dict[str, object]:
    customers = list(range(1, 16))
    metrics = held_karp_tsp(instance, customers)
    return {"method": "exact_held_karp", **route_summary(metrics)}


def solve_problem_2(instance: Instance) -> Dict[str, object]:
    customers = list(range(1, 16))
    metrics = exact_soft_time_window_tsp(instance, customers)
    return {"method": "exact_subset_time_dp", **route_summary(metrics)}


def solve_problem_3(instance: Instance) -> Dict[str, object]:
    metrics = best_single_vehicle_over_seeds(
        instance,
        instance.customers,
        seeds=range(10),
        extra_routes=None,
        max_rounds=3,
        enable_reverse=False,
    )
    return {"method": "multi_start_insertion_plus_local_search", **route_summary(metrics)}


def solve_fixed_vehicle(instance: Instance, vehicle_count: int) -> Dict[str, object]:
    initial_routes = build_fixed_vehicle_solution(instance, instance.customers, vehicle_count)
    improved_routes = improve_multi_vehicle_routes(instance, initial_routes, keep_nonempty=True, max_rounds=8)
    return {"vehicle_count": vehicle_count, **evaluate_routes(instance, improved_routes)}


def solve_problem_4(instance: Instance) -> Dict[str, object]:
    min_vehicle_lb = (sum(instance.demand[node] for node in instance.customers) + instance.capacity - 1) // instance.capacity
    best_primary = solve_fixed_vehicle(instance, min_vehicle_lb)
    analysis = [solve_fixed_vehicle(instance, vehicle_count) for vehicle_count in range(min_vehicle_lb, min_vehicle_lb + 4)]
    return {
        "method": "capacity_respecting_best_insertion_plus_route_relocation",
        "capacity": instance.capacity,
        "vehicle_lower_bound": min_vehicle_lb,
        "best_primary_solution": best_primary,
        "fixed_vehicle_analysis": analysis,
    }


def write_markdown_summary(path: Path, results: Dict[str, object]) -> None:
    lines = [
        "# MathorCup A Results Summary",
        "",
        "## Dataset",
        f"- Customer count: {results['dataset']['customer_count']}",
        f"- Vehicle capacity: {results['dataset']['capacity']}",
        f"- Total demand: {results['dataset']['total_demand']}",
        f"- Vehicle lower bound: {results['dataset']['vehicle_lower_bound']}",
    ]

    if "problem_1" in results:
        p1 = results["problem_1"]
        lines.extend(
            [
                "",
                "## Problem 1",
                f"- Method: {p1['method']}",
                f"- Route: {p1['route']}",
                f"- Travel time: {p1['travel_time']}",
            ]
        )

    if "problem_2" in results:
        p2 = results["problem_2"]
        lines.extend(
            [
                "",
                "## Problem 2",
                f"- Method: {p2['method']}",
                f"- Route: {p2['route']}",
                f"- Travel time: {p2['travel_time']}",
                f"- Time-window penalty: {p2['time_window_penalty']}",
                f"- Objective: {p2['objective']}",
            ]
        )

    if "problem_3" in results:
        p3 = results["problem_3"]
        lines.extend(
            [
                "",
                "## Problem 3",
                f"- Method: {p3['method']}",
                f"- Route length: {len(p3['route']) - 2}",
                f"- Travel time: {p3['travel_time']}",
                f"- Time-window penalty: {p3['time_window_penalty']}",
                f"- Objective: {p3['objective']}",
            ]
        )

    if "problem_4" in results:
        p4 = results["problem_4"]
        lines.extend(
            [
                "",
                "## Problem 4",
                f"- Method: {p4['method']}",
                f"- Vehicle lower bound: {p4['vehicle_lower_bound']}",
                f"- Best primary hierarchical objective: {p4['best_primary_solution']['hierarchical_objective']}",
                f"- Best primary secondary objective: {p4['best_primary_solution']['secondary_objective']}",
                "",
                "## Fixed Vehicle Analysis",
            ]
        )
        for item in p4["fixed_vehicle_analysis"]:
            lines.append(
                f"- {item['vehicle_count']} vehicles: travel {item['travel_time']}, penalty {item['time_window_penalty']}, secondary objective {item['secondary_objective']}"
            )

    path.write_text("\n".join(lines), encoding="utf-8")


def build_results(instance: Instance, selected_problems: Sequence[str]) -> Dict[str, object]:
    total_demand = sum(instance.demand[node] for node in instance.customers)
    vehicle_lower_bound = (total_demand + instance.capacity - 1) // instance.capacity

    results = {
        "dataset": {
            "customer_count": len(instance.customers),
            "capacity": instance.capacity,
            "total_demand": total_demand,
            "vehicle_lower_bound": vehicle_lower_bound,
        }
    }

    if "1" in selected_problems:
        results["problem_1"] = solve_problem_1(instance)
    if "2" in selected_problems:
        results["problem_2"] = solve_problem_2(instance)
    if "3" in selected_problems:
        results["problem_3"] = solve_problem_3(instance)
    if "4" in selected_problems:
        results["problem_4"] = solve_problem_4(instance)

    return results


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Solve the MathorCup 2026 A problem instance.")
    parser.add_argument("--excel", type=Path, default=None, help="Path to the reference workbook.")
    parser.add_argument("--json", type=Path, default=Path("results.json"), help="Path to write JSON results.")
    parser.add_argument("--markdown", type=Path, default=Path("results_summary.md"), help="Path to write Markdown summary.")
    parser.add_argument(
        "--problems",
        nargs="+",
        choices=["1", "2", "3", "4"],
        default=["1", "2", "3", "4"],
        help="Problem numbers to solve.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    excel_path = args.excel
    if excel_path is None:
        matches = sorted(Path(".").glob("*.xlsx"))
        if not matches:
            raise FileNotFoundError("No .xlsx workbook found in the current directory.")
        excel_path = matches[0]

    instance = load_instance(excel_path)
    results = build_results(instance, args.problems)
    args.json.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    write_markdown_summary(args.markdown, results)


if __name__ == "__main__":
    main()
